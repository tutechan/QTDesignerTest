#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Times   : 2019/11/21 8:55
# @Author  : zwj
# @Email   : zhangwj18@shanghai-electric.com
# @File    : __init__.py
# @Project: PyCharm

import requests
from openpyxl import Workbook
from PyQt5.QtWidgets import QApplication, QWidget, QMessageBox, QPushButton, QLineEdit, QLabel,QComboBox,QMainWindow
import sys
from dict_get import dict1,dict2,dict3
import urllib
import string
import datetime
from ChartGUI import Ui_MainWindow

class Mwindow(QMainWindow,Ui_MainWindow):
    def __init__(self):
        super(Mwindow,self).__init__()



        # self.lb1 = QLabel(self)
        # self.lb1.setGeometry(100, 200, 41, 31)
        # self.lb1.setText('品类：')
        # self.lb2 = QLabel(self)
        # self.lb2.setGeometry(100, 255, 41, 21)
        # self.lb2.setText('型号：')
        # self.lb3 = QLabel(self)
        # self.lb3.setGeometry(400, 200, 54, 21)
        # self.lb3.setText('开始时间：')
        # self.lb3 = QLabel(self)
        # self.lb3.setGeometry(400, 250, 54, 21)
        # self.lb3.setText('结束时间：')


        # self.cb1 = QComboBox(self)
        # self.cb1.setGeometry(140, 200, 200, 31)
        self.comboBox.addItems([x for x in dict1.keys()])
        self.comboBox.activated.connect(lambda: self.getData(self.comboBox.currentText()))

        # self.comboBox_2 = QComboBox(self)
        # self.comboBox_2.setGeometry(140, 250, 200, 31)
        #
        # self.textbox3 = QLineEdit(self)
        # self.textbox3.setGeometry(470, 200, 150, 31)
        # self.textbox4 = QLineEdit(self)
        # self.textbox4.setGeometry(470, 250, 150, 31)

        # self.findButton = QPushButton('数据抽取', self)
        # self.findButton.setGeometry(330, 330, 75, 33)
        # self.findButton.clicked.connect(self.getChartData)


    def getData(self,text):
        self.comboBox_2.clear()
        self.comboBox_2.addItems(dict1.get(text,""))
    #
    # def getChartData(self):
    #     url = "https://index.mysteel.com/newprice/getChartMultiCatalog.ms?callback=callback"
    #     # 请求头、参数设置
    #     headers = {
    #         "User-Agent": "PostmanRuntime/7.19.0",
    #         "Context-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    #         "Cookie": "Hm_lvt_1c4432afacfa2301369a5625795031b8=1572335579; href=https%3A%2F%2Findex.mysteel.com%2Fprice%2FindexPrice.html; accessId=5d36a9e0-919c-11e9-903c-ab24dbab411b; JSESSIONID=07332AA24253B7330274CD36B4312418; Hm_lpvt_1c4432afacfa2301369a5625795031b8=1572396557; qimo_seosource_5d36a9e0-919c-11e9-903c-ab24dbab411b=%E7%AB%99%E5%86%85; qimo_seokeywords_5d36a9e0-919c-11e9-903c-ab24dbab411b=; pageViewNum=4"}
    #
    #
    #     catalog_get = self.cb1.currentText()
    #     catalog_data = str(self.dict2.get(catalog_get,""))
    #     print(catalog_data)
    #     catalog = urllib.parse.quote(catalog_data, safe=string.printable)
    #     print(catalog)
    #
    #     spec_get = self.cb2.currentText()
    #     spec_data = str(self.dict3.get(spec_get,""))
    #     print(spec_data)
    #     spec = urllib.parse.quote(spec_data, safe=string.printable)
    #     print(spec)
    #
    #     #获取时间
    #     startTime = str(self.textbox3.text())
    #     endTime = str(self.textbox4.text())
    #     print(startTime, endTime)
    #
    #     data = {"catalog": catalog,
    #             "city": "%E4%B8%8A%E6%B5%B7",
    #             "spec": spec,
    #             "startTime": startTime,
    #             "endTime": endTime
    #             }
    #
    #     try:
    #         r= requests.post(url=url, data=data, headers=headers,verify=False)
    #         r.raise_for_status()
    #         r.encoding = 'utf-8'
    #         html = r.text
    #         print(html)
    #         print('成功')
    #
    #     except Exception as e:
    #         QMessageBox.information(self, "提示", "请确认输入是否正确！或网络是否为外网！/n"+str(e),
    #                                 QMessageBox.Yes)
    #     try:
    #     #网页解析
    #         chart_data = eval(html[9:-2])
    #         chart_list = chart_data['data']
    #         date = []
    #         value = []
    #         for k in chart_list:   #chart_list为一个列表
    #             lineNme = k['lineName']   #获得产品线
    #             # print(lineNme)
    #             for i in k['dateValueMap']:
    #                 date.append(i['date'])
    #                 value.append(i['value'])
    #
    #     # 写入excel
    #         wb = Workbook()  # 创建文件对象
    #         # grab the active worksheet
    #         ws = wb.active  # 获取第一个sheet
    #         ws.cell(1, 1).value = "Date"
    #         ws.cell(1, 2).value = "Price"
    #         for i in range(2,len(date)+2):
    #             ws.cell(i,1).value=datetime.datetime.strptime(date[i-2],"%Y-%m-%d").strftime("%Y/%m/%d")
    #
    #         for k in range(2,len(value)+2):
    #             ws.cell(k,2).value=('%.2f' % float(value[k-2]))
    #
    #         # wb.save(r"\\10.56.3.11\566服务中心\08_IT工具"+catalog_data+"-"+spec_data.replace("*"," ")+datetime.datetime.strptime(endTime,"%Y-%m-%d").strftime("%y%m%d")+".xlsx")
    #         wb.save("e:\\" + catalog_get + "-" + spec_get.replace("*", " ") + datetime.datetime.strptime(endTime, "%Y-%m-%d").strftime("%y%m%d") + ".xlsx")
    #         QMessageBox.information(self, "提示", "数据抽取成功！",
    #                                 QMessageBox.Yes)
    #
    #     except Exception as e:
    #         QMessageBox.information(self, "提示", "数据抽取失败！,请确认是否关闭EXCEL文件!/n"+str(e),
    #                                 QMessageBox.Yes)



if __name__ == '__main__':
    app = QApplication(sys.argv)
    find_weather = Mwindow()
    find_weather.show()
    sys.exit(app.exec())











